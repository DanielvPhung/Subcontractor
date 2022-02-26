from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import win32com.client
import time
from create_sc_form import create_sc_form
import os

#Get Important Files and respective paths
PATH = os.path.dirname(os.path.realpath(__file__))
FILE_NAME = 'Dummy Contact Info.xlsx'
PATH_TO_XLSX = (PATH + '\\' + FILE_NAME)

#xml column width, column widths are reduced by this error when saved
COLUMN_ERROR = 0.7109375

#In each sheet, we have 2 header rows with tabel labels
HEADER_ROWS = 2
#Used columns and respective widths per sheet
#Can guess this by respective string name and letter count, but used conservative values form trial and error
USED_COLUMNS  = ['A','B','C','D','E','F','G']
COLUMN_WIDTHS = [ 45, 12, 22, 14, 13, 27, 14]

def ensure_wb_closed():
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True
    wb = excel.Workbooks.Open(PATH_TO_XLSX)
    wb.Save()
    excel.Application.Quit()

def get_name():
    while(True):
        name = input('Enter Name: ')
        if(name == '0'):
            print('No name Entry')
            break
        name = name.title()
        return name

def get_sin():
    sin = ''
    while (True):
        sin = ''
        sin = input("Enter SIN: ")   
        sin = sin.replace(' ', '')
        if (sin == '0'):
            print("No SIN Entry")
            break
        elif (sin.isnumeric() == False):
            print("SIN Contains non-numeric elements")
        elif(len(sin) != 9):
            print("SIN does not contain 9 numbers")
        else:
            sin = (sin[0:3] + ' ' +
                   sin[3:6] + ' ' + 
                   sin[6:9])
            return sin
    return sin

def get_address():
    while(True):
        address = input('Enter Address: ')
        if(address == '0'):
            print('No address Entry')
            break
        address = address.title()
        return address

def get_city_province():
    while(True):
        city_province = input('Enter City and Province: ')
        if(city_province == '0'):
            print('No City and Province Entry')
            break
        return city_province

def get_postal_code():
    while(True):
        postal_code = input('Enter postal code: ')
        postal_code = postal_code.upper()
        postal_code = postal_code.replace(' ', '')
        if(postal_code == '0'):
            print('No Postal Code Entry')
            break
        if (len(postal_code) != 6):
            print('Postal Code  is not 6 digits')
        else:   
            postal_code = (postal_code[0:3] + ' ' + 
                           postal_code[3:6])
            return postal_code
        
def get_email():
    while(True):
        email = input('Enter Email: ')
        if(email == '0'):
            print('No email Entry')
            break
        return email

def get_phone_number():
    while(True):
        phone_number = input('Enter Phone Number: ')
        phone_number = phone_number.replace(' ', '')

        if(phone_number == '0'):
            print('No Phone Number Entry')
            break
        elif(phone_number.isnumeric() == False):
            print('Phone number has non-numeric values')
        elif (len(phone_number) != 10):
            print('Phone Number  is not 10 digits')
        else:
            phone_number = ( '(' + phone_number[0:3] + ') ' 
                                 + phone_number[3:6] + ' '
                                 +  phone_number[6:10] )
            return phone_number

def print_sc_list(ws):
    print()
    print(ws)
    i = HEADER_ROWS+1
    while(ws['A' + str(i)].value != None):
        print (str(i-2) + ': ' + ws['A' + str(i)].value)
        i+=1

def hire_new_sc(ws):
    new_row = ws.max_row+1
    #Assign
    while(True):
        name          = ws['A' + str(new_row)].value = get_name()
        sin           = ws['B' + str(new_row)].value = get_sin()
        address       = ws['C' + str(new_row)].value = get_address()
        city_province = ws['D' + str(new_row)].value = get_city_province()
        postal_code   = ws['E' + str(new_row)].value = get_postal_code()
        email         = ws['F' + str(new_row)].value = get_email()
        phone_number  = ws['G' + str(new_row)].value = get_phone_number()

        #print('Name is: ')
        print(' ' + ws['A' + str(new_row)].value)
        #print('SIN is: ')
        print(' ' + ws['B' + str(new_row)].value)
        #print('Address is: ')
        print(' ' + ws['C' + str(new_row)].value)
        #print('City, Province is: ')
        print(' ' + ws['D' + str(new_row)].value)
        #print('Postal Code is: ')
        print(' ' + ws['E' + str(new_row)].value)
        #print('Email is: ')
        print(' ' + ws['F' + str(new_row)].value)
        #print('Phone Number is: ')
        print(' ' + ws['G' + str(new_row)].value)

        if(input('Is Information this correct? Enter 1 to confirm, any other character to redo: ') == '1'):
            print('Entry Successfully Added.')
            print('Creating Subcontractor Form: ')
            create_sc_form(name,sin, address, city_province, postal_code, email, phone_number)
            break

def is_sheet_empty(sheet):
    return((sheet['A3'].value) == None)

def change_sc_status(og_sh, dest_sh):
    while(True):

        if (is_sheet_empty(og_sh)):
            print('Empty List! ')
            return False

        print_sc_list (og_sh)

        sc_row = input('Enter Corresponding Subcontractor Number (0 to quit): ')
        if(sc_row == '0'):
            return False
        else:
            sc_row = int(sc_row)+HEADER_ROWS
            print('Moving: ' + og_sh['A'+ str(sc_row)].value)

        name = og_sh['A'+ str(sc_row)].value
        save_entry = []

        for x in USED_COLUMNS:
            save_entry.append(og_sh[x + str(sc_row)].value)
        
        og_sh.delete_rows(sc_row)
        i=1
        new_row = dest_sh.max_row+1
        for x in USED_COLUMNS:
            dest_sh[x + str(new_row)].value = save_entry[i-1]
            i+=1

        print('Successfully moved ' + name)
        print()
        print('UPDATED SHEETS:')
        print_sc_list(og_sh)
        print_sc_list(dest_sh)
        return True
            
def format_widths(ws):
    i = 0
    for x in COLUMN_WIDTHS:
        ws.column_dimensions[USED_COLUMNS[i]].width = x + COLUMN_ERROR 
        i += 1

def format_borders(ws):
    top_left_cell  = 'A' + str(HEADER_ROWS+1)
    bot_right_cell = str(USED_COLUMNS[-1]) + str(ws.max_row)
    cell_range = top_left_cell + ':' + bot_right_cell


    side_border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                   )

    bottom_and_side_border = Border(bottom=Side(border_style='thin', color='000000' ),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000')
                            )

    all_border = Border(left   = Side(border_style='thin', color='000000'),
                        right  = Side(border_style='thin', color='000000'),
                        bottom = Side(border_style='thin', color='000000' ),
                        top    = Side(border_style='thin', color='000000' )
                       )

    no_bot_border = Border(bottom = Side(border_style=None))
    
    #If last row is empty, delete it
    #Caused by removing people
    if(ws['A'+ str(ws.max_row)].value == None):
        ws.delete_rows(ws.max_row)

    #Nested loop, column by column ->in every column go row by row
    for x in USED_COLUMNS:
        for i in range(HEADER_ROWS , ws.max_row+1):
            ws[x + str(i)].border = side_border
            
            if(i==HEADER_ROWS):
                ws[x + str(i)].border = all_border
            elif(i==ws.max_row):
                ws[x + str(i)].border = bottom_and_side_border
    
def sort_sheet_data(sheet_name):
    #win32com to change
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True

    wb = excel.Workbooks.Open(PATH_TO_XLSX)

    ws = wb.Worksheets(sheet_name)
    top_left_cell  = 'A' + str(HEADER_ROWS+1)
    bot_right_cell = str(USED_COLUMNS[-1]) + str(ws.UsedRange.Rows.Count)
    ws.Range(top_left_cell + ':' + bot_right_cell).Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
    wb.Save()
    excel.Application.Quit()

def main():

    #Make Sure Workbook is closed before altering
    ensure_wb_closed()

    #Workbook
    wb = load_workbook(PATH_TO_XLSX)

    #Load Sheet, get names of sheets
    curr_sh = wb.worksheets[0]
    prev_sh = wb.worksheets[1]
    name_curr_sc_sh = (curr_sh.title)
    name_prev_sc_sh = (prev_sh.title)

    #Continous Loop to gather input
    while (True):
        print('Enter 1 to Add New Subcontractor: ')
        print('Enter 2 to change Subcontractor Employment status: ')
        print('Enter 0 to quit program: ')

        status = input('Enter: ')
        match status:
            case '1':
                hire_new_sc(curr_sh)
                print('Sucessfully Added')

            case '2': 
                hire = input('Enter 1 to fire and 2 to rehire (any other key to quit): ')
                if(hire == '1'):
                    if(change_sc_status(curr_sh, prev_sh) == True):
                        print('Successfully Laid off')
                    else:
                        print('Nothing Done')

                elif(hire == '2'):
                    if(change_sc_status(prev_sh, curr_sh) == True):
                        print('Successfully Rehired')
                    else:
                        print('Nothing Done')              
                
            case '0':
                print()
                break

        print()

    #Save and Close
    format_widths(curr_sh)
    format_widths(prev_sh)

    format_borders(curr_sh)
    format_borders(prev_sh)
   
    wb.save(PATH_TO_XLSX)
    wb.close()

    print ('Reformatting Sheets....')
    
    #Sort Data On Excel
    sort_sheet_data(name_curr_sc_sh)
    sort_sheet_data(name_prev_sc_sh)
    print ('Reformatting Complete')
    print ('Sucessfully Exited Program')
    input('Press any key to close window....')

if __name__== "__main__" :
    main()





